from pathlib import Path
from datetime import datetime
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.contrib.auth.models import User, Group
from django.db import transaction

from openpyxl import load_workbook

from products.models import Product, Category, Brand, Supplier, Unit
from orders.models import PickupPoint, OrderStatus, Order, OrderItem


def _project_root() -> Path:
    p = Path(__file__).resolve()
    for parent in [p] + list(p.parents):
        if (parent / "manage.py").exists():
            return parent
    raise RuntimeError("Не найден корень проекта (manage.py)")


BASE_DIR = _project_root()
IMPORT_DIR = BASE_DIR / "import"


def _find_file(part: str) -> Path:
    part = part.lower()
    for p in IMPORT_DIR.glob("*.xlsx"):
        if part in p.name.lower():
            return p
    raise FileNotFoundError(f"Не найден xlsx по ключу '{part}' в папке {IMPORT_DIR}")


def _cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def _parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


class Command(BaseCommand):
    help = "Импорт данных из Excel"

    @transaction.atomic
    def handle(self, *args, **options):
        self.stdout.write(self.style.WARNING("Импорт начат..."))

        self.import_products()
        self.import_users()
        self.import_pickup_points()
        self.import_orders()

        self.stdout.write(self.style.SUCCESS("Импорт завершён ✅"))

    # =======================
    # ТОВАРЫ
    # =======================
    def import_products(self):
        f = _find_file("tovar")
        wb = load_workbook(f)
        ws = wb.active

        headers = [_cell_str(c.value) for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        need = [
            "Артикул",
            "Наименование товара",
            "Единица измерения",
            "Цена",
            "Поставщик",
            "Производитель",
            "Категория товара",
            "Действующая скидка",
            "Кол-во на складе",
        ]
        for n in need:
            if n not in col:
                raise ValueError(f"В {f.name} нет колонки: {n}. Нашёл: {headers}")

        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = _cell_str(row[col["Артикул"]])
            if not sku:
                continue

            name = _cell_str(row[col["Наименование товара"]])
            unit_name = _cell_str(row[col["Единица измерения"]])
            supplier_name = _cell_str(row[col["Поставщик"]])
            brand_name = _cell_str(row[col["Производитель"]])
            category_name = _cell_str(row[col["Категория товара"]])

            price = Decimal(str(row[col["Цена"]] or "0").replace(",", "."))
            discount = Decimal(str(row[col["Действующая скидка"]] or "0").replace(",", "."))
            qty = int(row[col["Кол-во на складе"]] or 0)

            category, _ = Category.objects.get_or_create(name=category_name)
            supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
            brand, _ = Brand.objects.get_or_create(name=brand_name)
            unit, _ = Unit.objects.get_or_create(name=unit_name, defaults={"abbreviation": ""})

            Product.objects.update_or_create(
                sku=sku,
                defaults=dict(
                    name=name,
                    category=category,
                    supplier=supplier,
                    brand=brand,
                    unit=unit,
                    price=price,
                    discount=discount,
                    quantity=qty,
                ),
            )

            created += 1

        self.stdout.write(self.style.SUCCESS(f"Товары: обработано {created}"))

    # =======================
    # ПОЛЬЗОВАТЕЛИ
    # =======================
    def import_users(self):
        f = _find_file("user")
        wb = load_workbook(f)
        ws = wb.active

        headers = [_cell_str(c.value) for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            role = _cell_str(row[col["Роль сотрудника"]])
            fio = _cell_str(row[col["ФИО"]])
            username = _cell_str(row[col["Логин"]])
            password = _cell_str(row[col["Пароль"]])

            if not username:
                continue

            user, was_created = User.objects.get_or_create(username=username)

            if password:
                user.set_password(password)

            parts = fio.split()
            user.last_name = parts[0] if len(parts) > 0 else ""
            user.first_name = parts[1] if len(parts) > 1 else ""
            user.save()

            if role:
                group, _ = Group.objects.get_or_create(name=role)
                user.groups.add(group)

            if was_created:
                created += 1

        self.stdout.write(self.style.SUCCESS(f"Пользователи: создано {created}"))

    # =======================
    # ПУНКТЫ ВЫДАЧИ
    # =======================
    def import_pickup_points(self):
        f = _find_file("пункт")
        wb = load_workbook(f)
        ws = wb.active

        created = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            addr = _cell_str(row[0])
            if not addr or "адрес" in addr.lower():
                continue

            _, was_created = PickupPoint.objects.get_or_create(address=addr)
            if was_created:
                created += 1

        self.stdout.write(self.style.SUCCESS(f"Пункты выдачи: создано {created}"))

    # =======================
    # ЗАКАЗЫ
    # =======================
    def import_orders(self):
        f = _find_file("заказ")
        wb = load_workbook(f)
        ws = wb.active

        headers = [_cell_str(c.value) for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            order_number = _cell_str(row[col["Номер заказа"]])
            items_raw = _cell_str(row[col["Артикул заказа"]])
            delivery_dt = _parse_date(row[col["Дата доставки"]])
            pickup_address = _cell_str(row[col["Адрес пункта выдачи"]])
            fio = _cell_str(row[col["ФИО авторизированного клиента"]])
            code = _cell_str(row[col["Код для получения"]])
            status_name = _cell_str(row[col["Статус заказа"]])

            if not order_number:
                continue

            status, _ = OrderStatus.objects.get_or_create(name=status_name)

            pickup, _ = PickupPoint.objects.get_or_create(address=pickup_address)

            username = fio.replace(" ", "_")[:150]
            customer, _ = User.objects.get_or_create(username=username)

            order, was_created = Order.objects.get_or_create(
                order_number=order_number,
                defaults=dict(
                    status=status,
                    pickup_point=pickup,
                    delivery_date=delivery_dt,
                    customer=customer,
                    pickup_code=code,   # <<< код для получения сохраняем сюда
                ),
            )

            if not was_created:
                order.status = status
                order.pickup_point = pickup
                order.delivery_date = delivery_dt
                order.customer = customer
                order.pickup_code = code
                order.save()

            parts = [p.strip() for p in items_raw.split(",") if p.strip()]
            for i in range(0, len(parts) - 1, 2):
                sku = parts[i]
                qty = int(parts[i + 1])

                product = Product.objects.filter(sku=sku).first()
                if not product:
                    continue

                OrderItem.objects.update_or_create(
                    order=order,
                    product=product,
                    defaults={"quantity": qty, "price": product.final_price},
                )

            if was_created:
                created += 1

        self.stdout.write(self.style.SUCCESS(f"Заказы: создано {created}"))